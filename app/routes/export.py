import csv
from io import StringIO, BytesIO
from decimal import Decimal

from datetime import datetime, time
from typing import Optional

from fastapi import APIRouter, Depends, Response, HTTPException, Query, Request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import select, func

from ..db import get_db
from ..models import Client, Order, Payment
from ..auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["export"])


def _fmt_dt(value: Optional[datetime]) -> str:
    if not value:
        return ""
    # Force Excel to treat as text to avoid #### when column is narrow.
    return "'" + value.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M")


def _fmt_money(value: Decimal) -> str:
    return f"{value:.2f}"


def _status_label(value: str) -> str:
    if value == "in_progress":
        return "In progress"
    return value.replace("_", " ").title()


def _xlsx_response(workbook: Workbook, filename: str) -> Response:
    buf = BytesIO()
    workbook.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _apply_header_style(ws, headers_count: int) -> None:
    header_font = Font(bold=True)
    for col in range(1, headers_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")


def _set_column_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _parse_date(value: Optional[str], field: str, is_end: bool = False) -> Optional[datetime]:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"{field} must be ISO date or datetime")
    if value and len(value) == 10:
        dt = datetime.combine(dt.date(), time.max if is_end else time.min)
    return dt


@router.get("/orders.csv")
def export_orders_csv(
    request: Request,
    client_id: Optional[int] = Query(None, ge=1),
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    get_current_user(request, db)
    tenant_id = getattr(getattr(request.state, "user", None), "tenant_id", 1)
    start_dt = _parse_date(date_from, "date_from", is_end=False)
    end_dt = _parse_date(date_to, "date_to", is_end=True)
    if status and status not in {"new", "in_progress", "done", "canceled"}:
        raise HTTPException(status_code=422, detail="status must be one of: new, in_progress, done, canceled")
    if start_dt and end_dt and start_dt > end_dt:
        raise HTTPException(status_code=422, detail="date_from must be <= date_to")

    # total paid per order
    paid_map = dict(
        db.execute(
            select(Payment.order_id, func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.tenant_id == tenant_id)
            .group_by(Payment.order_id)
        ).all()
    )

    stmt = select(Order, Client).join(Client, Client.id == Order.client_id)
    stmt = stmt.where(Order.tenant_id == tenant_id, Client.tenant_id == tenant_id)

    if client_id is not None:
        stmt = stmt.where(Order.client_id == client_id)
    if status:
        stmt = stmt.where(Order.status == status)
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where((Order.title.ilike(like)) | (Order.comment.ilike(like)))
    if start_dt:
        stmt = stmt.where(Order.created_at >= start_dt)
    if end_dt:
        stmt = stmt.where(Order.created_at <= end_dt)

    stmt = stmt.order_by(Order.id.desc())
    rows = db.execute(stmt).all()

    buf = StringIO()
    writer = csv.writer(buf, delimiter=";")

    writer.writerow([
        "Order ID",
        "Client ID",
        "Client Name",
        "Title",
        "Status",
        "Price",
        "Paid Total",
        "Balance",
        "Paid %",
        "Created At",
    ])

    for order, client in rows:
        paid_total = Decimal(str(paid_map.get(order.id, 0)))
        price = Decimal(str(order.price))
        balance = price - paid_total

        writer.writerow([
            order.id,
            client.id,
            client.name,
            order.title,
            _status_label(order.status),
            _fmt_money(price),
            _fmt_money(paid_total),
            _fmt_money(balance),
            f"{(paid_total / price * Decimal('100')):.2f}%" if price > 0 else "0.00%",
            _fmt_dt(order.created_at),
        ])

    content = "\ufeff" + buf.getvalue()
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="orders.csv"'},
    )


@router.get("/orders.xlsx")
def export_orders_xlsx(
    request: Request,
    client_id: Optional[int] = Query(None, ge=1),
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    get_current_user(request, db)
    tenant_id = getattr(getattr(request.state, "user", None), "tenant_id", 1)
    start_dt = _parse_date(date_from, "date_from", is_end=False)
    end_dt = _parse_date(date_to, "date_to", is_end=True)
    if status and status not in {"new", "in_progress", "done", "canceled"}:
        raise HTTPException(status_code=422, detail="status must be one of: new, in_progress, done, canceled")
    if start_dt and end_dt and start_dt > end_dt:
        raise HTTPException(status_code=422, detail="date_from must be <= date_to")

    paid_map = dict(
        db.execute(
            select(Payment.order_id, func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.tenant_id == tenant_id)
            .group_by(Payment.order_id)
        ).all()
    )

    stmt = select(Order, Client).join(Client, Client.id == Order.client_id)
    stmt = stmt.where(Order.tenant_id == tenant_id, Client.tenant_id == tenant_id)
    if client_id is not None:
        stmt = stmt.where(Order.client_id == client_id)
    if status:
        stmt = stmt.where(Order.status == status)
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where((Order.title.ilike(like)) | (Order.comment.ilike(like)))
    if start_dt:
        stmt = stmt.where(Order.created_at >= start_dt)
    if end_dt:
        stmt = stmt.where(Order.created_at <= end_dt)
    stmt = stmt.order_by(Order.id.desc())
    rows = db.execute(stmt).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = [
        "Order ID",
        "Client ID",
        "Client Name",
        "Title",
        "Status",
        "Price",
        "Paid Total",
        "Balance",
        "Paid %",
        "Created At",
    ]
    ws.append(headers)
    _apply_header_style(ws, len(headers))
    ws.freeze_panes = "A2"

    for order, client in rows:
        paid_total = Decimal(str(paid_map.get(order.id, 0)))
        price = Decimal(str(order.price))
        balance = price - paid_total
        paid_pct = float(paid_total / price) if price > 0 else 0.0
        ws.append([
            order.id,
            client.id,
            client.name,
            order.title,
            _status_label(order.status),
            float(price),
            float(paid_total),
            float(balance),
            paid_pct,
            order.created_at.replace(tzinfo=None) if order.created_at else None,
        ])

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = "0.00"
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.number_format = "0.00%"
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            cell.number_format = "yyyy-mm-dd hh:mm"

    _set_column_widths(ws, [10, 10, 24, 28, 14, 12, 12, 12, 10, 18])
    return _xlsx_response(wb, "orders.xlsx")


@router.get("/clients.csv")
def export_clients_csv(
    request: Request,
    ids: Optional[str] = Query(None, description="Comma-separated client ids"),
    db: Session = Depends(get_db),
):
    get_current_user(request, db)
    tenant_id = getattr(getattr(request.state, "user", None), "tenant_id", 1)
    stmt = select(Client).where(Client.tenant_id == tenant_id).order_by(Client.id.desc())
    if ids:
        try:
            id_list = [int(part) for part in ids.split(",") if part.strip()]
        except ValueError:
            raise HTTPException(status_code=422, detail="invalid ids")
        if id_list:
            stmt = stmt.where(Client.id.in_(id_list))
    clients = db.execute(stmt).scalars().all()

    buf = StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["Client ID", "Name", "Phone", "Telegram", "Notes", "Created At"])

    for c in clients:
        writer.writerow([
            c.id,
            c.name,
            c.phone or "",
            c.telegram or "",
            c.notes or "",
            _fmt_dt(c.created_at),
        ])

    content = "\ufeff" + buf.getvalue()
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="clients.csv"'},
    )


@router.get("/clients.xlsx")
def export_clients_xlsx(
    request: Request,
    ids: Optional[str] = Query(None, description="Comma-separated client ids"),
    db: Session = Depends(get_db),
):
    get_current_user(request, db)
    tenant_id = getattr(getattr(request.state, "user", None), "tenant_id", 1)
    stmt = select(Client).where(Client.tenant_id == tenant_id).order_by(Client.id.desc())
    if ids:
        try:
            id_list = [int(part) for part in ids.split(",") if part.strip()]
        except ValueError:
            raise HTTPException(status_code=422, detail="invalid ids")
        if id_list:
            stmt = stmt.where(Client.id.in_(id_list))
    clients = db.execute(stmt).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    headers = ["Client ID", "Name", "Phone", "Telegram", "Notes", "Created At"]
    ws.append(headers)
    _apply_header_style(ws, len(headers))
    ws.freeze_panes = "A2"

    for c in clients:
        ws.append([
            c.id,
            c.name,
            c.phone or "",
            c.telegram or "",
            c.notes or "",
            c.created_at.replace(tzinfo=None) if c.created_at else None,
        ])

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = "yyyy-mm-dd hh:mm"

    _set_column_widths(ws, [10, 24, 18, 18, 30, 18])
    return _xlsx_response(wb, "clients.xlsx")


@router.get("/templates/clients.csv")
def export_clients_template():
    buf = StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["name", "phone", "telegram", "notes"])
    content = "\ufeff" + buf.getvalue()
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="clients_template.csv"'},
    )


@router.get("/templates/orders.csv")
def export_orders_template():
    buf = StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["client_id", "title", "price", "status", "comment"])
    content = "\ufeff" + buf.getvalue()
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="orders_template.csv"'},
    )
